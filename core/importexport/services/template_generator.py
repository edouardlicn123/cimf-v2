# -*- coding: utf-8 -*-
"""
TemplateGenerator - 模板生成器

生成数据导入模板（Excel格式）
"""

from typing import Dict
from django.http import HttpResponse
from io import BytesIO

from core.importexport.services.import_service import ImportService


class TemplateGenerator:
    """导入模板生成器"""
    
    @classmethod
    def generate(cls, node_type_slug: str) -> HttpResponse:
        """生成导入模板"""
        from core.node.services import NodeTypeService
        
        node_type = NodeTypeService.get_by_slug(node_type_slug)
        if not node_type:
            raise ValueError(f"未找到节点类型: {node_type_slug}")
        
        fields = ImportService.get_importable_fields(node_type_slug)
        
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [f['label'] for f in fields]
        ws.append(headers)
        
        descriptions = [cls._get_description(f) for f in fields]
        ws.append(descriptions)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for cell in ws[2]:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 35
        
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass
            
            adjusted_width = min(max(max_length + 4, 15), 40)
            ws.column_dimensions[column].width = adjusted_width
        
        cls._add_fk_sheet(wb, node_type_slug)
        
        filename = f"{node_type_slug}_import_template.xlsx"
        return cls._build_response(wb, filename)
    
    @classmethod
    def _get_description(cls, field: Dict) -> str:
        """生成字段说明"""
        parts = []
        
        if field['required']:
            parts.append('必填')
        else:
            parts.append('可空')
        
        field_type = field['type']
        
        if field_type == 'fk':
            parts.append(f'FK参照，填写已有{field["label"]}名称')
        elif field_type == 'email':
            parts.append('邮箱格式')
        elif field_type == 'telephone':
            parts.append('文本格式')
        elif field_type == 'date':
            parts.append('格式：YYYY-MM-DD')
        elif field_type == 'datetime':
            parts.append('格式：YYYY-MM-DD HH:MM:SS')
        elif field_type == 'integer':
            parts.append('整数格式')
        elif field_type == 'decimal':
            parts.append('数字格式')
        elif field_type == 'json':
            from core.importexport.special_field_handler import SpecialFieldPool
            if SpecialFieldPool.is_special_field(field['name']):
                parts.append('格式：省份,城市,区县')
            else:
                max_len = field.get('max_length')
                if max_len:
                    parts.append(f'最大{max_len}字符')
        else:
            max_len = field.get('max_length')
            if max_len:
                parts.append(f'最大{max_len}字符')
        
        return '，'.join(parts)
    
    @classmethod
    def _add_fk_sheet(cls, wb, node_type_slug: str):
        """添加 FK 字段可选值工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws2 = wb.create_sheet("FK字段可选值")
        
        ws2.append(['FK字段名', '可选值'])
        
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 30
        
        fk_fields = ImportService.get_fk_fields_with_options(node_type_slug)
        
        MAX_ITEMS = 20
        
        for fk in fk_fields:
            items = fk['items']
            display_items = items[:MAX_ITEMS]
            
            if display_items:
                ws2.append([fk['label'], display_items[0]])
                for item in display_items[1:]:
                    ws2.append(['', item])
                
                if len(items) > MAX_ITEMS:
                    ws2.append(['', f'（其余{len(items) - MAX_ITEMS}个值请参考词汇表）'])
    
    @classmethod
    def _build_response(cls, wb, filename: str) -> HttpResponse:
        """构建 HTTP 响应"""
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response