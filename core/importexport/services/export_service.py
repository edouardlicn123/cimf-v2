# -*- coding: utf-8 -*-
"""
ExportService - 导出服务

提供通用的数据导出功能，支持 CSV/Excel 格式
支持自动发现模块字段，无需模块主动配置
"""

from typing import List, Dict, Any, Optional, Type
from django.http import HttpResponse
from django.db.models import Q
from datetime import datetime


class ExportService:
    """数据导出服务"""
    
    FORMAT_CSV = 'csv'
    FORMAT_XLSX = 'xlsx'
    
    EXPORTABLE_FIELD_TYPES = [
        'string', 'string_long', 'text', 'email', 'telephone',
        'integer', 'decimal', 'float', 'boolean',
        'entity_reference', 'datetime', 'date',
    ]
    
    FK_TYPES = ['fk', 'taxonomy', 'region']
    
    FILTERABLE_TYPES = ['string', 'string_long', 'text', 'email', 'telephone', 'fk']
    
    FILTERABLE_FK_FIELDS = ['country', 'industry', 'enterprise_nature', 'enterprise_type']
    
    @classmethod
    def get_exportable_fields(cls, node_type_slug: str) -> List[Dict]:
        """
        动态获取可导出的字段列表
        
        优先级：
        1. 模块配置的 export_fields（支持覆盖/补充/排除）
        2. 自动从 Django 模型发现
        """
        from core.importexport.model_registry import ModelRegistry
        from core.importexport.field_extractor import FieldDefExtractor
        
        module_config = cls._get_module_export_config(node_type_slug)
        
        if module_config is not None:
            auto_fields = cls._auto_discover_fields(node_type_slug, ModelRegistry, FieldDefExtractor)
            return FieldDefExtractor.merge_with_module_config(auto_fields, module_config)
        
        return cls._auto_discover_fields(node_type_slug, ModelRegistry, FieldDefExtractor)
    
    @classmethod
    def _auto_discover_fields(cls, node_type_slug: str, 
                              ModelRegistry, 
                              FieldDefExtractor) -> List[Dict]:
        """自动从 Django 模型发现字段"""
        model_class = ModelRegistry.get_model(node_type_slug)
        if model_class:
            return FieldDefExtractor.extract(model_class)
        return []
    
    @classmethod
    def _get_module_export_config(cls, node_type_slug: str) -> Optional[List[Dict]]:
        """获取模块配置的导出字段定义"""
        try:
            from importlib import import_module
            mod = import_module(f'modules.{node_type_slug}.module')
            if hasattr(mod, 'MODULE_INFO'):
                config = mod.MODULE_INFO.get('export_fields')
                if config:
                    return config
        except (ImportError, ModuleNotFoundError):
            pass
        return None
    
    @classmethod
    def get_fields_info(cls, node_type_slug: str, field_names: List[str]) -> List[Dict]:
        """获取选中字段的详细信息"""
        all_fields = cls.get_exportable_fields(node_type_slug)
        return [f for f in all_fields if f['name'] in field_names]
    
    @classmethod
    def get_filterable_fields(cls, node_type_slug: str) -> List[Dict]:
        """获取可筛选的字段列表"""
        all_fields = cls.get_exportable_fields(node_type_slug)
        return [f for f in all_fields if f['type'] in cls.FILTERABLE_TYPES]
    
    @classmethod
    def has_region_field(cls, node_type_slug: str) -> bool:
        """检查节点类型是否有省市区 JSON 字段"""
        all_fields = cls.get_exportable_fields(node_type_slug)
        return any(f['name'] == 'region' for f in all_fields)
    
    @classmethod
    def get_preview(cls, node_type_slug: str, field_names: List[str], 
                    filters: List[Dict] = None, limit: int = 5) -> List[Dict]:
        """获取数据预览"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        fields_info = cls.get_fields_info(node_type_slug, field_names)
        field_type_map = {f['name']: f['type'] for f in fields_info}
        
        preview_data = []
        for item in queryset[:limit]:
            row = cls._convert_to_row(item, field_names, field_type_map, node_type_slug)
            preview_data.append(row)
        
        return preview_data
    
    @classmethod
    def get_record_count(cls, node_type_slug: str, filters: List[Dict] = None) -> int:
        """获取记录总数"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        return queryset.count()
    
    @classmethod
    def export(cls, node_type_slug: str, field_names: List[str], 
               export_format: str = 'csv', filters: List[Dict] = None) -> HttpResponse:
        """执行导出"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        fields_info = cls.get_fields_info(node_type_slug, field_names)
        field_type_map = {f['name']: f['type'] for f in fields_info}
        
        rows = []
        for item in queryset:
            rows.append(cls._convert_to_row(item, field_names, field_type_map, node_type_slug))
        
        filename = cls.generate_filename(node_type_slug, export_format)
        
        if export_format == cls.FORMAT_CSV:
            return cls._export_csv(rows, fields_info, filename)
        else:
            return cls._export_xlsx(rows, fields_info, filename)
    
    @classmethod
    def generate_filename(cls, node_type_slug: str, export_format: str) -> str:
        """生成导出文件名"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f'{node_type_slug}_{timestamp}.{export_format}'
    
    @classmethod
    def _export_csv(cls, rows: List[Dict], fields: List[Dict], filename: str) -> HttpResponse:
        """导出为 CSV"""
        import csv
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([f['label'] for f in fields])
        
        for row in rows:
            writer.writerow([row.get(f['name'], '') for f in fields])
        
        return response
    
    @classmethod
    def _export_xlsx(cls, rows: List[Dict], fields: List[Dict], filename: str) -> HttpResponse:
        """导出为 XLSX"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [f['label'] for f in fields]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in rows:
            ws.append([row.get(f['name'], '') for f in fields])
        
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @classmethod
    def _convert_to_row(cls, item, field_names: List[str], 
                        field_type_map: Dict[str, str],
                        node_type_slug: str = None) -> Dict:
        """将数据对象转换为导出行"""
        row = {}
        
        for field_name in field_names:
            field_type = field_type_map.get(field_name, 'string')
            value = cls._get_field_value(item, field_name, field_type)
            row[field_name] = value
        
        return row
    
    @classmethod
    def _get_field_value(cls, obj: Any, field_name: str, field_type: str = 'string') -> Any:
        """获取字段值"""
        if obj is None:
            return ''
        
        if field_type in ['fk', 'taxonomy']:
            return cls._resolve_fk_field(obj, field_name)
        
        if field_type == 'region':
            return cls._resolve_region_field(obj)
        
        if field_type == 'boolean':
            value = getattr(obj, field_name, False)
            return '是' if value else '否'
        
        if field_type == 'datetime':
            value = getattr(obj, field_name, None)
            if value:
                return value.strftime('%Y-%m-%d %H:%M:%S')
            return ''
        
        if field_type == 'date':
            value = getattr(obj, field_name, None)
            if value:
                return value.strftime('%Y-%m-%d')
            return ''
        
        return getattr(obj, field_name, '') or ''
    
    @classmethod
    def _resolve_fk_field(cls, obj: Any, field_name: str) -> str:
        """解析 FK 字段，返回名称"""
        fk_obj = getattr(obj, field_name, None)
        if fk_obj is None:
            return ''
        
        if hasattr(fk_obj, 'name'):
            return fk_obj.name
        return str(fk_obj)
    
    @classmethod
    def _resolve_region_field(cls, obj: Any) -> str:
        """解析省市区 JSON 字段"""
        region = getattr(obj, 'region', None) or {}
        if isinstance(region, str):
            import json
            try:
                region = json.loads(region)
            except (json.JSONDecodeError, TypeError):
                return region
        
        province = region.get('province', '')
        city = region.get('city', '')
        district = region.get('district', '')
        parts = [p for p in [province, city, district] if p]
        return ' '.join(parts) if parts else ''
    
    @classmethod
    def _get_filtered_queryset(cls, node_type_slug: str, filters: List[Dict] = None):
        """获取应用筛选条件后的 QuerySet"""
        from core.importexport.model_registry import ModelRegistry
        
        model_class = ModelRegistry.get_model(node_type_slug)
        if not model_class:
            from core.node.models import Node
            queryset = Node.objects.filter(node_type__slug=node_type_slug)
            return cls._apply_filters(queryset, filters, node_type_slug, None, None)
        
        queryset = model_class.objects.all()
        
        if not filters:
            return queryset
        
        return cls._apply_filters(queryset, filters, node_type_slug, model_class)
    
    @classmethod
    def _apply_filters(cls, queryset, filters: List[Dict], 
                      node_type_slug: str, model_class: Type = None,
                      model_related_name: str = None):
        """应用筛选条件"""
        import json
        
        for f in filters:
            field = f.get('field', '')
            value = f.get('value', '')
            
            if not field or not value:
                continue
            
            if field == 'region':
                try:
                    region_data = json.loads(value) if isinstance(value, str) else value
                except (json.JSONDecodeError, TypeError):
                    region_data = {}
                q = Q()
                province = region_data.get('province', '')
                city = region_data.get('city', '')
                district = region_data.get('district', '')
                
                prefix = model_related_name or f'{node_type_slug}_fields'
                if province:
                    q &= Q(**{f'{prefix}__region__province__icontains': province})
                if city:
                    q &= Q(**{f'{prefix}__region__city__icontains': city})
                if district:
                    q &= Q(**{f'{prefix}__region__district__icontains': district})
                
                queryset = queryset.filter(q)
            elif model_class and hasattr(model_class, field):
                field_obj = model_class._meta.get_field(field)
                prefix = model_related_name or f'{node_type_slug}_fields'
                if isinstance(field_obj, models.ForeignKey):
                    queryset = queryset.filter(**{f'{prefix}__{field}__name__icontains': value})
                elif hasattr(field_obj, 'name') or field in ['charfield', 'textfield']:
                    queryset = queryset.filter(**{f'{prefix}__{field}__icontains': value})
        
        return queryset


from django.db import models
