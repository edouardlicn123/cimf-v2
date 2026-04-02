# -*- coding: utf-8 -*-
"""
================================================================================
文件：services.py
路径：/home/edo/cimf-v2/core/importexport/services.py
================================================================================

功能说明：
    数据导入导出服务，包含 ImportService, ExportService, TemplateGenerator
    
    主要功能：
    - 数据导入（CSV/Excel）
    - 数据导出（CSV/Excel）
    - 导入模板生成

版本：
    - 1.0: 从 modules/services/ 迁移合并

依赖：
    - core.node.models: Node, NodeType
    - core.node.services: NodeTypeService
    - nodes.customer.services: CustomerService
    - nodes.customer_cn.services: CustomerCnService
"""

import re
import io
from typing import List, Dict, Any, Optional, Tuple
from django.http import HttpResponse
from django.db.models import Q
from datetime import datetime


# ========================
# ImportService - 导入服务
# ========================

class ImportService:
    """
    通用导入服务
    
    设计思路：
    - 通用方法，适用于所有节点类型
    - 通过 ModelRegistry 动态获取模型
    - 通过 FKResolverPool 和 SpecialFieldPool 处理特殊字段
    """
    
    FORMAT_CSV = 'csv'
    FORMAT_XLSX = 'xlsx'
    
    FK_TAXONOMY_OVERRIDES = {
        ('customer_cn', 'enterprise_type'): 'enterprise_nature',
    }
    
    @classmethod
    def get_importable_fields(cls, node_type_slug: str) -> List[Dict]:
        """获取可导入的字段列表（通用方法）"""
        from core.importexport.model_registry import ModelRegistry
        from core.importexport.field_extractor import FieldDefExtractor
        
        model_class = ModelRegistry.get_model(node_type_slug)
        if not model_class:
            return []
        
        return FieldDefExtractor.extract(model_class)
    
    @classmethod
    def read_file(cls, file, format: str) -> Tuple[List[str], List[List[str]]]:
        """读取文件内容"""
        if format == cls.FORMAT_CSV:
            return cls._read_csv(file)
        else:
            return cls._read_xlsx(file)
    
    @classmethod
    def _read_csv(cls, file) -> Tuple[List[str], List[List[str]]]:
        """读取 CSV 文件"""
        import csv
        
        decoded_file = file.read().decode('utf-8-sig')
        reader = csv.reader(decoded_file.splitlines())
        rows = list(reader)
        
        if not rows:
            return [], []
        
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        
        return headers, data_rows
    
    @classmethod
    def _read_xlsx(cls, file) -> Tuple[List[str], List[List[str]]]:
        """读取 XLSX 文件"""
        from openpyxl import load_workbook
        
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        
        rows = list(ws.values)
        
        if not rows:
            return [], []
        
        headers = [str(h) if h is not None else '' for h in rows[0]]
        data_rows = [[str(cell) if cell is not None else '' for cell in row] for row in rows[1:]]
        
        return headers, data_rows
    
    @classmethod
    def map_headers_to_fields(cls, headers: List[str], fields: List[Dict]) -> Dict[str, str]:
        """将文件头部映射到字段定义"""
        header_to_field = {}
        header_lower_map = {h.lower(): h for h in headers}
        
        for field in fields:
            field_label = field['label'].lower()
            
            if field_label in header_lower_map:
                header_to_field[header_lower_map[field_label]] = field['name']
            else:
                field_name = field['name'].lower()
                if field_name in header_lower_map:
                    header_to_field[header_lower_map[field_name]] = field['name']
        
        return header_to_field
    
    @classmethod
    def parse_data(cls, headers: List[str], data_rows: List[List[str]], 
                   header_to_field: Dict[str, str]) -> List[Dict]:
        """解析数据行"""
        parsed_rows = []
        
        for row in data_rows:
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    header = headers[i]
                    if header in header_to_field:
                        field_name = header_to_field[header]
                        row_dict[field_name] = str(cell).strip() if cell else ''
            
            parsed_rows.append(row_dict)
        
        return parsed_rows
    
    @classmethod
    def validate_data(cls, node_type_slug: str, rows: List[Dict]) -> Dict:
        """验证数据"""
        fields = cls.get_importable_fields(node_type_slug)
        field_map = {f['name']: f for f in fields}
        
        valid_count = 0
        errors = []
        
        for idx, row in enumerate(rows, start=1):
            row_errors = []
            
            for field_name, value in row.items():
                if field_name not in field_map:
                    continue
                
                field = field_map[field_name]
                field_errors = cls._validate_field(field, value)
                row_errors.extend(field_errors)
            
            if row_errors:
                errors.append({
                    'row': idx,
                    'data': row,
                    'errors': row_errors,
                })
            else:
                valid_count += 1
        
        return {
            'valid_count': valid_count,
            'error_count': len(errors),
            'errors': errors,
        }
    
    @classmethod
    def _validate_field(cls, field: Dict, value: Any) -> List[str]:
        """验证单个字段"""
        errors = []
        
        if not value:
            if field['required']:
                errors.append(f"{field['label']} 不能为空")
            return errors
        
        field_type = field['type']
        
        if field_type == 'email':
            if not cls._is_valid_email(value):
                errors.append(f"{field['label']} 邮箱格式不正确")
        
        elif field_type == 'fk':
            from core.importexport.fk_resolver import FKResolverPool
            fk_model = field.get('fk_model')
            if fk_model:
                resolved = FKResolverPool.resolve(fk_model, value)
                if not resolved:
                    errors.append(f"{field['label']} '{value}' 不存在")
        
        elif field_type == 'json':
            from core.importexport.special_field_handler import SpecialFieldPool
            if SpecialFieldPool.is_special_field(field['name']):
                pass
        
        return errors
    
    @classmethod
    def _is_valid_email(cls, email: str) -> bool:
        """验证邮箱格式"""
        pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        return bool(re.match(pattern, str(email)))
    
    @classmethod
    def import_data(cls, node_type_slug: str, rows: List[Dict], 
                    user, skip_duplicates: bool = True) -> Dict:
        """执行导入"""
        from core.importexport.model_registry import ModelRegistry
        from core.importexport.fk_resolver import FKResolverPool
        from core.importexport.special_field_handler import SpecialFieldPool
        from core.node.models import Node, NodeType
        
        model_class = ModelRegistry.get_model(node_type_slug)
        fields = cls.get_importable_fields(node_type_slug)
        field_map = {f['name']: f for f in fields}
        
        success_count = 0
        errors = []
        
        node_type = NodeType.objects.filter(slug=node_type_slug).first()
        if not node_type:
            raise ValueError(f"未找到节点类型: {node_type_slug}")
        
        for idx, row in enumerate(rows, start=1):
            try:
                transformed = cls._transform_row(row, node_type_slug, field_map)
                
                existing = cls._find_existing(model_class, transformed)
                
                if existing:
                    if skip_duplicates:
                        continue
                    instance = existing
                else:
                    node = Node.objects.create(
                        node_type=node_type,
                        created_by=user,
                        updated_by=user,
                    )
                    instance = model_class.objects.create(node=node)
                
                for key, value in transformed.items():
                    setattr(instance, key, value)
                
                instance.save()
                success_count += 1
                
            except Exception as e:
                errors.append({
                    'row': idx,
                    'data': row,
                    'errors': [str(e)],
                })
        
        return {
            'success_count': success_count,
            'error_count': len(errors),
            'errors': errors,
        }
    
    @classmethod
    def _transform_row(cls, row: Dict, node_type_slug: str, field_map: Dict) -> Dict:
        """转换行数据"""
        from core.importexport.fk_resolver import FKResolverPool
        from core.importexport.special_field_handler import SpecialFieldPool
        
        transformed = {}
        
        for field_name, value in row.items():
            if field_name not in field_map:
                continue
            
            field = field_map[field_name]
            field_type = field['type']
            
            if not value:
                continue
            
            if field_type == 'fk':
                fk_model = field.get('fk_model')
                if fk_model:
                    taxonomy_slug = cls.FK_TAXONOMY_OVERRIDES.get(
                        (node_type_slug, field_name),
                        field_name
                    )
                    resolved = FKResolverPool.resolve(fk_model, value, taxonomy_slug)
                    if resolved is not None:
                        transformed[field_name] = resolved
            
            elif field_type == 'json':
                if SpecialFieldPool.is_special_field(field_name):
                    transformed[field_name] = SpecialFieldPool.handle_import(field_name, value)
                else:
                    transformed[field_name] = value
            
            else:
                transformed[field_name] = value
        
        return transformed
    
    @classmethod
    def _find_existing(cls, model_class, data: Dict):
        """查找已存在的记录"""
        for field_name, value in data.items():
            try:
                field = model_class._meta.get_field(field_name)
            except Exception:
                continue
            if getattr(field, 'unique', False) and value:
                existing = model_class.objects.filter(**{field_name: value}).first()
                if existing:
                    return existing
        return None
    
    @classmethod
    def get_fk_fields_with_options(cls, node_type_slug: str) -> List[Dict]:
        """获取 FK 字段及其可选值"""
        from core.importexport.model_registry import ModelRegistry
        from core.models import Taxonomy, TaxonomyItem
        
        fields = cls.get_importable_fields(node_type_slug)
        
        result = []
        
        for field in fields:
            if field['type'] != 'fk':
                continue
            
            field_name = field['name']
            
            taxonomy_slug = cls.FK_TAXONOMY_OVERRIDES.get(
                (node_type_slug, field_name), 
                field_name
            )
            
            taxonomy = Taxonomy.objects.filter(slug=taxonomy_slug).first()
            
            if taxonomy:
                items = list(TaxonomyItem.objects.filter(
                    taxonomy=taxonomy
                ).values_list('name', flat=True).order_by('weight', 'name'))
                
                result.append({
                    'name': field_name,
                    'label': field['label'],
                    'items': items,
                    'total': len(items)
                })
        
        return result
    
    @classmethod
    def generate_error_csv(cls, errors: List[Dict], fields: List[Dict]) -> HttpResponse:
        """生成错误列表 CSV"""
        import csv
        
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="import_errors.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['行号', '错误原因', '数据'])
        
        for error in errors:
            row_num = error.get('row', '')
            error_msgs = '; '.join(error.get('errors', []))
            data = str(error.get('data', ''))
            writer.writerow([row_num, error_msgs, data])
        
        return response


# ========================
# ExportService - 导出服务
# ========================

class ExportService:
    """数据导出服务"""
    
    FORMAT_CSV = 'csv'
    FORMAT_XLSX = 'xlsx'
    
    EXPORTABLE_FIELD_TYPES = [
        'string', 'string_long', 'text', 'email', 'telephone',
        'integer', 'decimal', 'float', 'boolean',
        'entity_reference', 'datetime', 'date',
    ]
    
    # 从字段类型判断是否需要特殊处理
    FK_TYPES = ['fk', 'taxonomy', 'region']
    
    FILTERABLE_TYPES = ['string', 'string_long', 'text', 'email', 'telephone', 'fk']
    
    @classmethod
    def get_exportable_fields(cls, node_type_slug: str) -> List[Dict]:
        """动态获取可导出的字段列表"""
        from importlib import import_module
        
        try:
            mod = import_module(f'modules.{node_type_slug}.module')
            if hasattr(mod, 'MODULE_INFO'):
                return mod.MODULE_INFO.get('export_fields', [])
        except (ImportError, ModuleNotFoundError):
            pass
        
        return []
    
    @classmethod
    def get_fields_info(cls, node_type_slug: str, field_names: List[str]) -> List[Dict]:
        """获取选中字段的详细信息"""
        all_fields = cls.get_exportable_fields(node_type_slug)
        return [f for f in all_fields if f['name'] in field_names]
    
    @classmethod
    def get_filterable_fields(cls, node_type_slug: str) -> List[Dict]:
        """获取可筛选的字段列表"""
        all_fields = cls.get_exportable_fields(node_type_slug)
        return [f for f in all_fields if f['type'] in cls.FILTERABLE_TYPES]
    
    @classmethod
    def has_region_field(cls, node_type_slug: str) -> bool:
        """检查节点类型是否有省市区字段"""
        all_fields = cls.get_exportable_fields(node_type_slug)
        return any(f['name'] == 'region' for f in all_fields)
    
    @classmethod
    def get_preview(cls, node_type_slug: str, field_names: List[str], 
                    filters: List[Dict] = None, limit: int = 5) -> List[Dict]:
        """获取数据预览"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        fields = cls.get_fields_info(node_type_slug, field_names)
        
        preview_data = []
        for item in queryset[:limit]:
            row = cls._convert_to_row(item, field_names, node_type_slug)
            preview_data.append(row)
        
        return preview_data
    
    @classmethod
    def get_record_count(cls, node_type_slug: str, filters: List[Dict] = None) -> int:
        """获取记录总数"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        return queryset.count()
    
    @classmethod
    def export(cls, node_type_slug: str, field_names: List[str], 
               export_format: str = 'csv', filters: List[Dict] = None) -> HttpResponse:
        """执行导出"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        fields = cls.get_fields_info(node_type_slug, field_names)
        
        rows = []
        for item in queryset:
            rows.append(cls._convert_to_row(item, field_names, node_type_slug))
        
        filename = cls.generate_filename(node_type_slug, export_format)
        
        if export_format == cls.FORMAT_CSV:
            return cls._export_csv(rows, fields, filename)
        else:
            return cls._export_xlsx(rows, fields, filename)
    
    @classmethod
    def generate_filename(cls, node_type_slug: str, export_format: str) -> str:
        """生成导出文件名"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f'{node_type_slug}_{timestamp}.{export_format}'
    
    @classmethod
    def _export_csv(cls, rows: List[Dict], fields: List[Dict], filename: str) -> HttpResponse:
        """导出为 CSV"""
        import csv
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([f['label'] for f in fields])
        
        for row in rows:
            writer.writerow([row.get(f['name'], '') for f in fields])
        
        return response
    
    @classmethod
    def _export_xlsx(cls, rows: List[Dict], fields: List[Dict], filename: str) -> HttpResponse:
        """导出为 XLSX"""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [f['label'] for f in fields]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row in rows:
            ws.append([row.get(f['name'], '') for f in fields])
        
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @classmethod
    def _convert_to_row(cls, item, field_names: List[str], 
                        node_type_slug: str = None) -> Dict:
        """将数据对象转换为导出行"""
        # 获取字段类型映射
        fields = cls.get_exportable_fields(node_type_slug) if node_type_slug else []
        field_type_map = {f['name']: f['type'] for f in fields}
        
        row = {}
        for field_name in field_names:
            field_type = field_type_map.get(field_name, 'string')
            value = cls._get_field_value(item, field_name, field_type)
            row[field_name] = value
        return row
    
    @classmethod
    def _get_field_value(cls, item, field_name: str, field_type: str = 'string') -> Any:
        """获取字段值"""
        if field_type in ['fk', 'taxonomy']:
            return cls._resolve_fk_field(item, field_name)
        
        if field_type == 'region':
            return cls._resolve_region_field(item)
        
        if field_type == 'boolean':
            value = getattr(item, field_name, False)
            return '是' if value else '否'
        
        return getattr(item, field_name, '') or ''
    
    @classmethod
    def _resolve_fk_field(cls, item, field_name: str) -> str:
        """解析 FK 字段，返回名称"""
        obj = getattr(item, field_name, None)
        return obj.name if obj and hasattr(obj, 'name') else ''
    
    @classmethod
    def _resolve_region_field(cls, item) -> str:
        """解析省市区 JSON 字段"""
        region = getattr(item, 'region', None) or {}
        province = region.get('province', '')
        city = region.get('city', '')
        district = region.get('district', '')
        parts = [p for p in [province, city, district] if p]
        return ' '.join(parts) if parts else ''
    
    @classmethod
    def _get_service_class(cls, node_type_slug: str):
        """动态获取服务类"""
        try:
            from importlib import import_module
            mod = import_module(f'modules.{node_type_slug}.services')
            
            # 查找 Service 结尾的类
            for attr_name in dir(mod):
                attr = getattr(mod, attr_name)
                if (attr_name.endswith('Service') and 
                    hasattr(attr, 'get_list')):
                    return attr
        except (ImportError, ModuleNotFoundError):
            pass
        
        return None
    
    @classmethod
    def _get_filtered_queryset(cls, node_type_slug: str, filters: List[Dict] = None):
        """获取应用筛选条件后的 QuerySet"""
        service = cls._get_service_class(node_type_slug)
        if not service:
            raise ValueError(f"Unknown node type: {node_type_slug}")
        
        queryset = service.get_list()
        
        if not filters:
            return queryset
        
        for f in filters:
            field = f.get('field', '')
            value = f.get('value', '')
            
            if not field or not value:
                continue
            
            if field == 'region':
                import json
                try:
                    region_data = json.loads(value) if isinstance(value, str) else value
                except (json.JSONDecodeError, TypeError):
                    region_data = {}
                province = region_data.get('province', '')
                city = region_data.get('city', '')
                district = region_data.get('district', '')
                q = Q()
                if province:
                    q &= Q(**{f'region__province__icontains': province})
                if city:
                    q &= Q(**{f'region__city__icontains': city})
                if district:
                    q &= Q(**{f'region__district__icontains': district})
                queryset = queryset.filter(q)
            elif field in cls.FILTERABLE_FK_FIELDS:
                queryset = queryset.filter(**{f'{field}__name__icontains': value})
            elif field in cls.FILTERABLE_FIELD_TYPES:
                queryset = queryset.filter(**{f'{field}__icontains': value})
        
        return queryset


# ========================
# TemplateGenerator - 模板生成器
# ========================

class TemplateGenerator:
    """导入模板生成器"""
    
    @classmethod
    def generate(cls, node_type_slug: str) -> HttpResponse:
        """生成导入模板"""
        from core.node.services import NodeTypeService
        
        node_type = NodeTypeService.get_by_slug(node_type_slug)
        if not node_type:
            raise ValueError(f"未找到节点类型: {node_type_slug}")
        
        fields = ImportService.get_importable_fields(node_type_slug)
        
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [f['label'] for f in fields]
        ws.append(headers)
        
        descriptions = [cls._get_description(f) for f in fields]
        ws.append(descriptions)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for cell in ws[2]:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 35
        
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass
            
            adjusted_width = min(max(max_length + 4, 15), 40)
            ws.column_dimensions[column].width = adjusted_width
        
        cls._add_fk_sheet(wb, node_type_slug)
        
        filename = f"{node_type_slug}_import_template.xlsx"
        return cls._build_response(wb, filename)
    
    @classmethod
    def _get_description(cls, field: Dict) -> str:
        """生成字段说明"""
        parts = []
        
        if field['required']:
            parts.append('必填')
        else:
            parts.append('可空')
        
        field_type = field['type']
        
        if field_type == 'fk':
            parts.append(f'FK参照，填写已有{field["label"]}名称')
        elif field_type == 'email':
            parts.append('邮箱格式')
        elif field_type == 'telephone':
            parts.append('文本格式')
        elif field_type == 'date':
            parts.append('格式：YYYY-MM-DD')
        elif field_type == 'datetime':
            parts.append('格式：YYYY-MM-DD HH:MM:SS')
        elif field_type == 'integer':
            parts.append('整数格式')
        elif field_type == 'decimal':
            parts.append('数字格式')
        elif field_type == 'json':
            from core.importexport.special_field_handler import SpecialFieldPool
            if SpecialFieldPool.is_special_field(field['name']):
                parts.append('格式：省份,城市,区县')
            else:
                max_len = field.get('max_length')
                if max_len:
                    parts.append(f'最大{max_len}字符')
        else:
            max_len = field.get('max_length')
            if max_len:
                parts.append(f'最大{max_len}字符')
        
        return '，'.join(parts)
    
    @classmethod
    def _add_fk_sheet(cls, wb, node_type_slug: str):
        """添加 FK 字段可选值工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws2 = wb.create_sheet("FK字段可选值")
        
        ws2.append(['FK字段名', '可选值'])
        
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 30
        
        fk_fields = ImportService.get_fk_fields_with_options(node_type_slug)
        
        MAX_ITEMS = 20
        
        for fk in fk_fields:
            items = fk['items']
            display_items = items[:MAX_ITEMS]
            
            if display_items:
                ws2.append([fk['label'], display_items[0]])
                for item in display_items[1:]:
                    ws2.append(['', item])
                
                if len(items) > MAX_ITEMS:
                    ws2.append(['', f'（其余{len(items) - MAX_ITEMS}个值请参考词汇表）'])
    
    @classmethod
    def _build_response(cls, wb, filename: str) -> HttpResponse:
        """构建 HTTP 响应"""
        from io import BytesIO
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response